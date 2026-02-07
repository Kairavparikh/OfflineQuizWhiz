"""
CSV/Excel exporter for exam papers.

Exports papers to CSV format matching client's template:
- Test Section
- Main Topic
- Sub-topic
- Difficulty Level
- Question ID
- Question (English)
- Option A
- Option B
- Option C
- Option D
- Correct Answer
- Explanation
- References
"""

import csv
from typing import List
from pathlib import Path
from paper_builder import Paper
from models import Question


def export_paper_to_csv(paper: Paper, output_path: str) -> str:
    """
    Export a paper to CSV format matching client's template.

    Args:
        paper: Paper object to export
        output_path: Path to output CSV file

    Returns:
        Path to created CSV file
    """
    output_file = Path(output_path)

    # Define column headers matching client's exact template
    headers = [
        "Test Section",
        "Main Topic",
        "Sub-topic",
        "Difficulty Level",
        "Translation for options required?",
        "Question ID",
        "Question- English",
        "Question- Hindi",
        "Option A- English",
        "Option A- Hindi",
        "Option B- English",
        "Option B- Hindi",
        "Option C- English",
        "Option C- Hindi",
        "Option D- English",
        "Option D- Hindi",
        "Correct Answer",
        "Solution/Workout/Explanation",
        "Reference(s)"
    ]

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Write header
        writer.writerow(headers)

        # Write each question
        for q in paper.questions:
            # Format references as numbered list
            if q.references:
                references_str = "\n".join([f"{i}. {ref}" for i, ref in enumerate(q.references, 1)])
            else:
                references_str = ""

            # Format correct answer as "Option X"
            correct_answer_formatted = f"Option {q.correct_answer}"

            row = [
                q.test_section,
                q.main_topic,
                q.subtopic,
                q.difficulty.value,
                "",  # Translation for options required? (empty for now)
                q.question_id,
                q.question_text_en,
                "",  # Question- Hindi (empty for now)
                q.option_a_en,
                "",  # Option A- Hindi (empty for now)
                q.option_b_en,
                "",  # Option B- Hindi (empty for now)
                q.option_c_en,
                "",  # Option C- Hindi (empty for now)
                q.option_d_en,
                "",  # Option D- Hindi (empty for now)
                correct_answer_formatted,
                q.explanation,
                references_str
            ]

            writer.writerow(row)

    print(f"\n✅ Exported {len(paper.questions)} questions to: {output_file}")
    print(f"   File size: {output_file.stat().st_size / 1024:.1f} KB")

    return str(output_file)


def export_questions_to_csv(questions: List[Question], output_path: str) -> str:
    """
    Export a list of questions to CSV format matching client's template.

    Args:
        questions: List of Question objects
        output_path: Path to output CSV file

    Returns:
        Path to created CSV file
    """
    output_file = Path(output_path)

    # Define column headers matching client's exact template
    headers = [
        "Test Section",
        "Main Topic",
        "Sub-topic",
        "Difficulty Level",
        "Translation for options required?",
        "Question ID",
        "Question- English",
        "Question- Hindi",
        "Option A- English",
        "Option A- Hindi",
        "Option B- English",
        "Option B- Hindi",
        "Option C- English",
        "Option C- Hindi",
        "Option D- English",
        "Option D- Hindi",
        "Correct Answer",
        "Solution/Workout/Explanation",
        "Reference(s)"
    ]

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Write header
        writer.writerow(headers)

        # Write each question
        for q in questions:
            # Format references as numbered list
            if q.references:
                references_str = "\n".join([f"{i}. {ref}" for i, ref in enumerate(q.references, 1)])
            else:
                references_str = ""

            # Format correct answer as "Option X"
            correct_answer_formatted = f"Option {q.correct_answer}"

            row = [
                q.test_section,
                q.main_topic,
                q.subtopic,
                q.difficulty.value,
                "",  # Translation for options required? (empty for now)
                q.question_id,
                q.question_text_en,
                "",  # Question- Hindi (empty for now)
                q.option_a_en,
                "",  # Option A- Hindi (empty for now)
                q.option_b_en,
                "",  # Option B- Hindi (empty for now)
                q.option_c_en,
                "",  # Option C- Hindi (empty for now)
                q.option_d_en,
                "",  # Option D- Hindi (empty for now)
                correct_answer_formatted,
                q.explanation,
                references_str
            ]

            writer.writerow(row)

    print(f"\n✅ Exported {len(questions)} questions to: {output_file}")
    print(f"   File size: {output_file.stat().st_size / 1024:.1f} KB")

    return str(output_file)


def export_paper_to_excel(paper: Paper, output_path: str) -> str:
    """
    Export a paper to Excel format.

    Requires openpyxl: pip install openpyxl

    Args:
        paper: Paper object to export
        output_path: Path to output Excel file

    Returns:
        Path to created Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("⚠️  openpyxl not installed. Install with: pip install openpyxl")
        print("   Falling back to CSV export...")
        csv_path = output_path.replace('.xlsx', '.csv')
        return export_paper_to_csv(paper, csv_path)

    output_file = Path(output_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    # Define headers
    headers = [
        "Test Section",
        "Main Topic",
        "Sub-topic",
        "Difficulty Level",
        "Question ID",
        "Question (English)",
        "Option A",
        "Option B",
        "Option C",
        "Option D",
        "Correct Answer",
        "Explanation",
        "References"
    ]

    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write questions
    for row_num, q in enumerate(paper.questions, 2):
        references_str = "; ".join(q.references) if q.references else "N/A"

        row_data = [
            q.test_section,
            q.main_topic,
            q.subtopic,
            q.difficulty.value,
            q.question_id,
            q.question_text_en,
            q.option_a_en,
            q.option_b_en,
            q.option_c_en,
            q.option_d_en,
            q.correct_answer,
            q.explanation,
            references_str
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value

            # Wrap text for long columns
            if col_num in [6, 7, 8, 9, 10, 12]:  # Question, options, explanation
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        if col_num in [1, 2, 3, 4, 5, 11]:  # Metadata columns
            ws.column_dimensions[chr(64 + col_num)].width = 15
        elif col_num in [6, 7, 8, 9, 10]:  # Question and options
            ws.column_dimensions[chr(64 + col_num)].width = 40
        elif col_num == 12:  # Explanation
            ws.column_dimensions[chr(64 + col_num)].width = 50
        else:  # References
            ws.column_dimensions[chr(64 + col_num)].width = 30

    # Save workbook
    wb.save(output_file)

    print(f"\n✅ Exported {len(paper.questions)} questions to: {output_file}")
    print(f"   File size: {output_file.stat().st_size / 1024:.1f} KB")

    return str(output_file)


# Example usage
if __name__ == "__main__":
    from models import Question, DifficultyLevel
    from paper_builder import Paper
    from datetime import datetime
    import uuid

    # Create sample questions
    questions = [
        Question(
            question_id=str(uuid.uuid4()),
            test_section="Main Subject",
            main_topic="Material Science",
            subtopic="Crystal Structure",
            difficulty=DifficultyLevel.EASY,
            question_text_en="What is the coordination number in an FCC crystal structure?",
            option_a_en="12",
            option_b_en="8",
            option_c_en="6",
            option_d_en="4",
            correct_answer="A",
            explanation="In Face-Centered Cubic (FCC) structure, each atom is surrounded by 12 nearest neighbors, giving a coordination number of 12.",
            references=["Callister, Materials Science and Engineering, Chapter 3"]
        ),
        Question(
            question_id=str(uuid.uuid4()),
            test_section="Main Subject",
            main_topic="Thermodynamics",
            subtopic="Phase Diagrams",
            difficulty=DifficultyLevel.MEDIUM,
            question_text_en="At what temperature does the eutectoid transformation occur in the Fe-C phase diagram?",
            option_a_en="727°C",
            option_b_en="912°C",
            option_c_en="1147°C",
            option_d_en="1538°C",
            correct_answer="A",
            explanation="The eutectoid transformation in the Fe-C system occurs at 727°C (0.8% C), where austenite transforms to pearlite.",
            references=["ASM Handbook, Volume 3: Alloy Phase Diagrams"]
        )
    ]

    # Create sample paper
    paper = Paper(
        paper_id=str(uuid.uuid4()),
        paper_name="Sample Exam 2026",
        subject="Metallurgical Engineering",
        questions=questions,
        created_at=datetime.now().isoformat()
    )

    # Export to CSV
    print("\n" + "="*80)
    print("TESTING CSV EXPORT")
    print("="*80)
    export_paper_to_csv(paper, "sample_paper.csv")

    # Export to Excel
    print("\n" + "="*80)
    print("TESTING EXCEL EXPORT")
    print("="*80)
    export_paper_to_excel(paper, "sample_paper.xlsx")

    print("\n" + "="*80)
    print("✅ Export tests complete!")
    print("="*80)
    print("\nCheck current directory for:")
    print("  - sample_paper.csv")
    print("  - sample_paper.xlsx")
